# scores_ophalen.py
# Exporteert data/<SEIZOEN_STARTJAAR>/<Maand>.csv vanuit Whistloting.xlsm
# - Maand uit Loting!H8 (enige datumbron)
# - Seizoen-map kies je zelf (CLI --jaar of prompt, met suggestie op basis van seizoenlogica)
#
# Vereist:
#   pip install pandas openpyxl

from __future__ import annotations

import argparse
import os
import shutil
import sys
import tempfile
import time
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Any

try:
    import pandas as pd
except Exception:
    print("Pandas ontbreekt. Installeer met: pip install pandas")
    sys.exit(1)

try:
    import openpyxl  # noqa: F401
except Exception:
    print("openpyxl ontbreekt. Installeer met: pip install openpyxl")
    sys.exit(1)

SHEET_DATA = "samen"
SHEET_DATUM = "Loting"
CELL_H8_0BASED = (7, 7)  # H8 -> (row=7, col=7) 0-based
MAX_RIJEN = 55

MAANDEN_12 = [
    "Januari", "Februari", "Maart", "April", "Mei", "Juni",
    "Juli", "Augustus", "September", "Oktober", "November", "December"
]


# -------------------------
# Helpers
# -------------------------
def project_root() -> Path:
    return Path(__file__).resolve().parent


def is_excel_lockfile(p: Path) -> bool:
    # Excel creëert lock files zoals "~$Whistloting.xlsm"
    return p.name.startswith("~$")


def vriendelijke_lock_melding(xlsm_path: Path) -> None:
    folder = xlsm_path.parent
    lock = folder / f"~${xlsm_path.name}"

    print("\n[FOUT] Toegang geweigerd tot het Excel-bestand.")
    print("Waarschijnlijk staat Whistloting.xlsm nog open in Excel (of er is een lockfile actief).")
    print(f"Bestand: {xlsm_path}")
    if lock.exists():
        print(f"Lockfile gevonden: {lock}")
    print("\n➡️  Sluit Excel volledig (alle vensters), wacht 2 seconden en probeer opnieuw.\n")


def vind_xlsm(explicit: Optional[str] = None) -> Path:
    """
    Zoek Whistloting.xlsm (voorkeur) of whis*.xlsm in projectroot of 1 map hoger.
    Negeert Excel lockfiles (~$...).
    """
    if explicit:
        p = Path(explicit).expanduser().resolve()
        if not p.is_file():
            print(f"[XLSM] Opgegeven bestand niet gevonden: {p}")
            sys.exit(1)
        if is_excel_lockfile(p):
            print(f"[XLSM] Opgegeven bestand is een Excel-lockfile en kan niet gelezen worden: {p}")
            vriendelijke_lock_melding(p)
            sys.exit(1)
        return p

    roots = [project_root(), project_root().parent]

    # 1) Prefer exact Whistloting.xlsm (case-insensitive)
    for root in roots:
        p = root / "Whistloting.xlsm"
        if p.is_file() and not is_excel_lockfile(p):
            return p
        # ook case-insensitive check
        try:
            for candidate in root.glob("*.xlsm"):
                if candidate.is_file() and candidate.name.lower() == "whistloting.xlsm" and not is_excel_lockfile(candidate):
                    return candidate
        except Exception:
            pass

    # 2) Fallback: whis*.xlsm (most recent)
    kandidaten: list[Path] = []
    for root in roots:
        try:
            for p in root.glob("*.xlsm"):
                if not p.is_file():
                    continue
                if is_excel_lockfile(p):
                    continue
                name = p.name.lower()
                if "whis" in name:  # jouw oorspronkelijke heuristiek
                    kandidaten.append(p)
        except Exception:
            pass

    if not kandidaten:
        print(f"[XLSM] Geen Whistloting.xlsm/whis*.xlsm gevonden in {roots}.")
        print("Zet je Excelbestand in de projectmap of gebruik --xlsm PAD.")
        sys.exit(1)

    kandidaten.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return kandidaten[0]


def read_excel_safe(xlsm_path: Path, **read_kwargs) -> pd.DataFrame:
    """
    Lees Excel via pandas/openpyxl.
    Als bestand vergrendeld is: toon duidelijke melding en stop (geen traceback).
    """
    # Extra safeguard: nooit lockfile lezen
    if is_excel_lockfile(xlsm_path):
        print(f"[XLSM] Lockfile gedetecteerd: {xlsm_path}")
        vriendelijke_lock_melding(xlsm_path)
        sys.exit(1)

    try:
        return pd.read_excel(str(xlsm_path), engine="openpyxl", **read_kwargs)
    except PermissionError:
        # Duidelijke boodschap i.p.v. copy-truc: gebruiker moet Excel sluiten.
        vriendelijke_lock_melding(xlsm_path)
        sys.exit(1)


def parse_excel_date(value: Any) -> pd.Timestamp:
    """
    Converteer H8-waarde naar pandas Timestamp.
    Ondersteunt datetime/date, Excel-serial (int/float), en strings.
    """
    if isinstance(value, (datetime, date)):
        return pd.to_datetime(value)

    if isinstance(value, (int, float)):
        # Excel serial date (met Excel origin 1899-12-30)
        ts = pd.to_datetime(value, unit="D", origin="1899-12-30", errors="coerce")
        if not pd.isna(ts):
            return ts

    if isinstance(value, str):
        v = value.strip()
        if not v:
            raise ValueError("Datumcel H8 is leeg.")
        ts = pd.to_datetime(v, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            ts = pd.to_datetime(v, errors="coerce")
        if not pd.isna(ts):
            return ts

    raise ValueError(f"Onbekende datumwaarde in {SHEET_DATUM}!H8: {value!r}")


def haal_datum_uit_loting(xlsm_path: Path) -> pd.Timestamp:
    """
    Lees datum uit Loting!H8.
    Eerst via pandas (snel), dan openpyxl fallback (maar ook met PermissionError-melding).
    """
    # 1) pandas route
    try:
        df = read_excel_safe(xlsm_path, sheet_name=SHEET_DATUM, header=None)
        val = df.iloc[CELL_H8_0BASED[0], CELL_H8_0BASED[1]]
        return parse_excel_date(val)
    except SystemExit:
        raise
    except Exception:
        pass

    # 2) openpyxl fallback
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsm_path), data_only=True, read_only=True)
        ws = wb[SHEET_DATUM]
        val = ws["H8"].value
        return parse_excel_date(val)
    except PermissionError:
        vriendelijke_lock_melding(xlsm_path)
        sys.exit(1)


def seizoen_startjaar(ts: pd.Timestamp) -> int:
    """Seizoen start in september: Sep–Dec => ts.year, Jan–Aug => ts.year - 1."""
    return int(ts.year) if ts.month >= 9 else int(ts.year) - 1


def output_pad(seizoen: str, maand: str) -> Path:
    out_dir = project_root() / "data" / seizoen
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir / f"{maand}.csv"


# -------------------------
# Main
# -------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Exporteer CSV uit Whistloting.xlsm")
    parser.add_argument("--jaar", help="Seizoen-map (startjaar, bv. 2025)")
    parser.add_argument("--xlsm", help="Pad naar Excel .xlsm")
    args = parser.parse_args()

    xlsm = vind_xlsm(args.xlsm)
    print(f"[XLSM] Bestand: {xlsm}")

    ts = haal_datum_uit_loting(xlsm)
    maandnaam = MAANDEN_12[int(ts.month) - 1]
    suggestie = seizoen_startjaar(ts)

    print(f"[DATA] Loting!H8 = {ts.date()} -> maand: {maandnaam}  (seizoen-suggestie: {suggestie})")

    seizoen = args.jaar
    if not seizoen:
        try:
            inp = input(f"Seizoen-map (startjaar, bv. {suggestie}) [{suggestie}]: ").strip()
        except EOFError:
            inp = ""
        seizoen = inp or str(suggestie)

    if not seizoen.isdigit():
        print(f"[CLI] Ongeldige seizoen-map: {seizoen!r} (verwacht bv. 2025)")
        sys.exit(1)

    # Enkel waarschuwingen (we schrijven altijd naar gekozen map)
    if ts.month >= 9 and int(seizoen) != ts.year:
        print(f"[WAARSCHUWING] Datum is {maandnaam} {ts.year}, maar je schrijft naar map {seizoen}.")
    if ts.month < 9 and int(seizoen) != ts.year - 1:
        print(f"[WAARSCHUWING] Datum is {maandnaam} {ts.year}, maar volgens seizoenlogica is map {ts.year - 1}.")

    # Data uit 'samen'
    df = read_excel_safe(xlsm, sheet_name=SHEET_DATA)
    print(f"[XLSM] Sheet '{SHEET_DATA}' ingeladen. Vorm: {df.shape}")

    csv_path = output_pad(seizoen, maandnaam)
    df.iloc[:MAX_RIJEN, :].to_csv(csv_path, index=False, encoding="utf-8")
    print(f"[OK] Geschreven: {csv_path}")


if __name__ == "__main__":
    main()
